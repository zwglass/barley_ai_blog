from pathlib import Path
import os
# from openpyxl import Workbook
from openpyxl import load_workbook, Workbook
import pandas as pd
# pip install pandas openpyxl


class ZwExcelHandler(object):
    """
    from zwutils_methods import ZwExcelHandler        # excel文件基础操作
    #    self.cls_excel_handle = ZwExcelHandler(filepath)          # excel文件基础操作
    """
    def __init__(self, filename=None):
        self.filename = filename

    def column_letter_to_number(self, letter: str) -> int:
        """
        将 Excel 列字母（如 "A"、"AA"）转换为列数字（如 1、27）
        
        Args:
            letter (str): 列字母（不区分大小写），如 "A"、"B"、"Z"、"AA"、"BC"。
        
        Returns:
            int: 对应的列数字（从 1 开始）。
        
        Example:
            >>> column_letter_to_number("A")  # 返回 1
            >>> column_letter_to_number("Z")  # 返回 26
            >>> column_letter_to_number("AA") # 返回 27
        """
        letter = letter.upper()  # 确保大写
        num = 0
        for char in letter:
            num = num * 26 + (ord(char) - ord('A') + 1)
        return num


    def column_number_to_letter(self, number: int) -> str:
        """
        将列数字（如 1、27）转换为 Excel 列字母（如 "A"、"AA"）
        
        Args:
            number (int): 列数字（≥1）。
        
        Returns:
            str: 对应的列字母。
        
        Example:
            >>> column_number_to_letter(1)   # 返回 "A"
            >>> column_number_to_letter(26) # 返回 "Z"
            >>> column_number_to_letter(27) # 返回 "AA"
        """
        if number < 1:
            raise ValueError("列数字必须 ≥1")
        
        letter = ""
        while number > 0:
            number -= 1
            letter = chr(ord('A') + (number % 26)) + letter
            number = number // 26
        return letter

    def read_excel_workbook(self, excel_path):
        # 读取 excel 返回 Workbook
        wb = load_workbook(excel_path)
        return wb
    
    def workbook_write(self, wb, write_path):
        # 保存 workbook
        wb.save(write_path)
    
    def ws_column_compute_max_row_num(self, ws, column_letter) -> int:
        # 计算列最大的有数据的行; 行从 1 开始
        # 获取 column_letter 列最后一个有数据的行号
        last_row = ws.max_row  # 先获取最大可能行号

        column_num = self.column_letter_to_number(column_letter)
        # 从下往上找，直到遇到非空单元格
        while last_row > 0 and ws.cell(row=last_row, column=column_num).value is None:
            last_row -= 1
        return last_row

    def ws_update_cell_value(self, ws, row_number, column_letter, new_value):
        # 更新ws表制定单元格的值
        column_num = self.column_letter_to_number(column_letter)
        ws.cell(row=row_number, column=column_num, value=new_value)
        return ws
    
    def ws_read_cell_value(self, ws, row_number, column_letter):
        # 读取单元格值
        column_num = self.column_letter_to_number(column_letter)
        v = ws.cell(row=row_number, column=column_num).value
        return v

    def excel_insert_row(self, excel_path, row_num = 1):
        # excel 插入行
        # 加载 workbook
        wb = load_workbook(excel_path)
        ws = wb.active  # ws 为活跃的 worksheet

        # 插入顶部行
        ws.insert_rows(row_num)  # 将第 row_num 从 1开始; 1插入到顶部

        # 保存更改
        wb.save(excel_path)

    def series_insert_excel(self, excel_path, insert_start_row, insert_column_num, insert_series: pd.Series):
        """ 插入行 """
        # excel 插入行
        # 加载 workbook
        wb = load_workbook(excel_path)
        ws = wb.active  # ws 为活跃的 worksheet

        # 将 series 插入到指定列（以 'F' 为示例）
        for i, value in enumerate(insert_series):
            ws.cell(row=i+insert_start_row, column=insert_column_num).value = value  # 索引从 0 开始，所以 row 和 column 需要 +1

        # 保存更改
        wb.save(excel_path)
    
    def to_pandas(self, excel_path, sheet_name=None, header=0):
        """ 从Excel文件读取数据到pandas.DataFrame; header 是从 0 开始, None 没有header 会用0开始的序号代替 """
        try:
            if sheet_name:
                return pd.read_excel(excel_path, sheet_name=sheet_name, header=header)
            else:
                return pd.read_excel(excel_path, header=header)
        except Exception as e:
            print(e)
            return None
    
    def pandas_write_to_excel(self, save_excel_path, df, sheet_name=None):
        """ 将pandas.DataFrame写入Excel文件 save_excel_path: .xlsx 文件路径   (不能是 .xls)"""
        if sheet_name is None:
            df.to_excel(save_excel_path, index=False)  # index=False 表示不保存索引列
        else:
            df.to_excel(save_excel_path, index=False, sheet_name=sheet_name)

        # 多个 pandas 保存到同一个 excel
        # with pd.ExcelWriter(save_excel_path, engine='openpyxl') as writer:
        #     df1.to_excel(writer, sheet_name='Sheet1', index=False)
        #     df2.to_excel(writer, sheet_name='Sheet2', index=False)

    def df_rename_columns(self, df: pd.DataFrame, rename_columns: dict):
        # pandas.DataFrame 重命名 列名称 rename_columns = {'col1': 'New Col 1', ...}
        renamed = df.rename(columns=rename_columns)
        return renamed
    

    def df_loc_columns(self, df: pd.DataFrame, columns: list):
        # pandas根据列名保留数据 筛选列名 columns = ['col1', 'col2']
        return df[columns]
    
    def df_to_list_dict(self, df: pd.DataFrame):
        # pandas DataFrame 转 list
        return df.to_dict(orient='records')
    
    def series_columns(self, series: pd.Series, columns: list):
        # pandas根据列名筛选数据 筛选列名 columns = ['col1', 'col2']
        s1 = pd.Series(series, index=columns)
        return s1
    
    def df_columns_sort(self, df, sort_columns_list):
        # pandas 按列的顺序自定义排序
        sorted_df = df[sort_columns_list]
        return sorted_df
    
    def df_add_columns(self, df, new_columns):
        # pandas 添加列; new_columns = {'C': '', 'D': ''}  # 多列相同值
        df = df.assign(**new_columns)
        return df

    def unmerge_excel_cells(self, file_path, output_path=None):
        """
        取消 Excel 文件中的所有合并单元格，并将合并区域的值填充到每个单元格。
        
        参数:
            file_path (str): 输入的 Excel 文件路径。
            output_path (str): 输出的 Excel 文件路径。如果未提供，将覆盖原文件。
        
        返回:
            None
        """
        # 加载 Excel 文件
        workbook = load_workbook(file_path)
        sheet = workbook.active

        # 遍历所有合并单元格区域
        for merged_cell in list(sheet.merged_cells.ranges):  # 使用列表防止遍历时修改
            # 获取合并单元格区域的左上角单元格值
            top_left_cell = sheet.cell(row=merged_cell.min_row, column=merged_cell.min_col)
            value = top_left_cell.value

            # 取消合并单元格
            sheet.unmerge_cells(str(merged_cell))

            # 将值填充到合并区域的每个单元格
            for row in range(merged_cell.min_row, merged_cell.max_row + 1):
                for col in range(merged_cell.min_col, merged_cell.max_col + 1):
                    sheet.cell(row=row, column=col, value=value)

        # 保存结果
        unmerge_file_path = Path(file_path).parent / f"nomerge_{Path(file_path).name}"
        output_path = output_path or str(unmerge_file_path)
        workbook.save(output_path)
        print(f"已取消合并单元格，结果保存为: {output_path}")
        return output_path

    def delete_columns_by_conditions(self, file_path, row_value_list, output_path = None):
        """
        删除 Excel 文件中满足指定多行条件的列。
        参数:
            file_path (str): 输入的 Excel 文件路径。
            output_path (str): 输出的 Excel 文件路径。
            row_value_list (list): 包含字典的列表，每个字典定义行号和需要匹配的值。
                示例: [{'row': 2, 'value': 'row_2_value'}, {'row': 3, 'value': 'row_3_val}, ...]
                row: start 1;

        返回:
            None
        """
        # 加载 Excel 文件
        workbook = load_workbook(file_path)
        sheet = workbook.active

        # 记录需要删除的列索引
        delete_columns = []

        # 遍历所有列
        for col in range(1, sheet.max_column + 1):
            # 检查该列是否满足所有条件
            match = True
            for condition in row_value_list:
                row = condition['row']
                expected_value = condition['value']
                # 获取对应行的单元格值
                cell_value = sheet.cell(row=row, column=col).value
                # 如果不满足条件，标记为 False 并退出
                if not cell_value or not cell_value.startswith(expected_value):
                    match = False
                    break
            # 如果满足所有条件，记录列索引
            if match:
                delete_columns.append(col)

        # 删除记录的列，从最后一列开始删除，避免索引错乱
        for col in sorted(delete_columns, reverse=True):
            sheet.delete_cols(col)


        # 保存结果
        output_path = output_path or file_path
        workbook.save(output_path)
        print(f"已删除指定列，结果保存为: {output_path}")
        return output_path
    
    def delete_rows_by_indices(self, file_path, rows_to_delete, output_path=None):
        """
        删除 Excel 文件中的指定行。

        参数:
            file_path (str): 输入的 Excel 文件路径。
            output_path (str): 输出的 Excel 文件路径。
            rows_to_delete (list): 需要删除的行索引列表（从 1 开始）。

        返回:
            None
        """
        # 加载 Excel 文件
        workbook = load_workbook(file_path)
        sheet = workbook.active

        # 删除指定行，从后往前删除避免索引错乱
        for row_index in sorted(rows_to_delete, reverse=True):
            sheet.delete_rows(row_index)

        # 保存结果
        output_path = output_path or file_path
        workbook.save(output_path)
        print(f"已删除指定行，结果保存为: {output_path}")
        return output_path
    
    def delete_rows_by_conditions(self, file_path, columns_value_list, max_rows = 0, output_path = None):
        """
        删除 Excel 文件中满足指定多列条件的行。
        当满足条件的行数 > max_rows 时，不再删除。
        参数:
            file_path (str): 输入的 Excel 文件路径。
            output_path (str): 输出的 Excel 文件路径。
            columns_value_list (list): 包含字典的列表，每个字典定义行号和需要匹配的值。
                示例: [{'column': 2, 'value': 'column_2_val'}, {'column': 3, 'value': 'column_3_val}, ...]
                row: start 1;

        返回:
            output_path
        """
        # 检查文件是否存在
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"The file {file_path} does not exist.")

        # 读取 Excel 文件
        df = pd.read_excel(file_path, header=None)

        # 转换列号从 1 开始到 0 开始（Pandas 使用 0 索引）
        conditions = []
        for cond in columns_value_list:
            column_index = cond['column'] - 1
            column_name = df.columns[column_index]  # 获取列名
            value = cond['value']
            conditions.append((column_name, value))

        # 找到满足条件的行索引
        rows_to_delete = pd.Series([True] * len(df))
        for column_name, value in conditions:
            # rows_to_delete &= (df[column_name] == value)
            rows_to_delete &= df[column_index].astype(str).str.startswith(str(value))


        # 计算满足条件的行数
        matching_rows_count = rows_to_delete.sum()

        # 如果满足条件的行数超过 max_rows，则停止删除
        if max_rows > 0 and matching_rows_count > max_rows:
            print(f"Number of matching rows ({matching_rows_count}) exceeds max_rows ({max_rows}). No rows deleted.")
            return file_path

        # 删除满足条件的行
        df = df[~rows_to_delete]

        # 设置默认输出路径
        if output_path is None:
            output_path = file_path

        # 保存结果到 Excel 文件
        df.to_excel(output_path, index=False, header=False)  # 保留无列名的结构

        return output_path
    
    def merge_excel_files(self, folder_path, output_file=None):
        """
        将文件夹内的所有 Excel 文件合并为一个文件。

        参数:
            folder_path (str): 包含 Excel 文件的文件夹路径。
            output_file (str): 合并后生成的 Excel 文件路径。

        返回:
            None
        """
        # 创建一个新的工作簿用于保存合并后的数据
        df = self.read_excel_files(folder_path)
        foder_path_excel_file = Path(folder_path).parent / f"merged_{Path(folder_path).name}.xlsx"
        output_file = output_file or str(foder_path_excel_file)
        self.pandas_write_to_excel(output_file, df)
        return output_file

    def read_excel_files(self, folder_path, sheet_name=0, combine=True):
        """
        读取文件夹内所有 Excel 文件，并加载为 Pandas 数据框。

        参数:
            folder_path (str): 包含 Excel 文件的文件夹路径。
            sheet_name (str or int or None): 要读取的工作表名称或索引，默认为第一个工作表 (0)。
            combine (bool): 如果为 True，将所有数据框合并为一个，默认为 True。
        
        返回:
            list 或 DataFrame: 如果 combine=False，返回数据框列表；如果 combine=True，返回一个合并后的数据框。
        """
        # 获取文件夹内所有 Excel 文件
        excel_files = [f for f in os.listdir(folder_path) if f.endswith(('.xls', '.xlsx'))]

        if not excel_files:
            print("No Excel files found in the specified folder.")
            return []

        dataframes = []

        for file in excel_files:
            file_path = os.path.join(folder_path, file)
            try:
                # 读取 Excel 文件
                df = pd.read_excel(file_path, sheet_name=sheet_name)
                dataframes.append(df)
            except Exception as e:
                print(f"Failed to read {file}: {e}")

        if combine:
            # 合并所有数据框
            return pd.concat(dataframes, ignore_index=True)
        else:
            return dataframes

    
    def handle_excel_files(self, folder_path, handle_function):
        # 文件夹内的所有 Excel 文件统一操作
        for filename in os.listdir(folder_path):
            if filename.endswith(".xlsx") or filename.endswith(".xlsm"):  # 检查文件扩展名
                file_path = os.path.join(folder_path, filename)
                # print(f"正在处理文件: {filename}")
                if handle_function:
                    handle_function(file_path)
                else:
                    print(f"未指定处理函数，跳过文件: {filename}")

    def unique_column_value(self, pd_column: pd.Series):
        # pandas列去重
        return pd_column.drop_duplicates()
    
    def df_write_to_exist_excel(self, excel_path:str, sheet_name:str, df_data:pd.DataFrame) -> str:
        # 把 DataFrame 数据追加到已经存在的excel文件内
        with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            book = writer.book
            if sheet_name in book.sheetnames:
                startrow = book[sheet_name].max_row
            else:
                startrow = 0
            
            df_data.to_excel(
                writer, 
                sheet_name=sheet_name, 
                startrow=startrow, 
                index=False, 
                header=False if startrow > 0 else True
            )
        return excel_path

# 使用示例
if __name__ == "__main__":
    # 创建并写入数据
    # run: python zwutils_methods/excel_handle.py
    excel_path = '/Users/senmalay/Downloads/My WangWang/0105 11-42-05 8280export.xlsx'
    cls_excel_handle = ZwExcelHandler(excel_path)
    # df = cls_excel_handle.to_pandas()
    excel_path_unmerged = cls_excel_handle.unmerge_excel_cells(excel_path)

    # 删除符合条件的列
    row_value_list = [
        {'row': 2, 'value': 'SKU信息'},
        {'row': 3, 'value': '商家编码'}
    ]
    excel_path_delete_columns = cls_excel_handle.delete_columns_by_conditions(excel_path_unmerged, row_value_list)

    # 删除指定的行
    rows_to_delete = [1, 2]
    excel_path_delete_rows = cls_excel_handle.delete_rows_by_indices(excel_path_delete_columns, rows_to_delete)
    
    # # 把列名称改为英文
    # en_columns_obj = {'宝贝id': 'goods_id_taobao', '宝贝标题': 'name', '一口价': 'price'}
    # renamed_df = cls_excel_handle.df_rename_columns(df, en_columns_obj)

    # # 重新保存为excel
    # from pathlib import Path
    # new_excel_name = f'test_renamed_{Path(excel_path).stem}.xlsx'
    # new_excel_path = str(Path(excel_path).parent / new_excel_name)
    # cls_excel_handle.pandas_write_to_excel(new_excel_path, renamed_df, None)
    # print('renamed_excel path:', new_excel_path)
